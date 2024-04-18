from typing import List
from pandas import read_excel
from datetime import datetime
from copy import deepcopy
import random
import datetime
import numpy as np
from openpyxl import Workbook

def ConcatFilePathFileName(filepath, filename):
    if filepath[len(filepath) - 1] != '\\':
        filepath += '\\'
    filepath += filename
    return filepath

def prepareArrays():

    bus_levels=[]
    bus_TODs=[]
    bus_names=[]
    bus_A_phase_voltages=[]
    bus_B_phase_voltages=[]
    bus_C_phase_voltages=[]
    bus_N_voltages=[]
    bus_A_phase_voltage_angles=[]
    bus_B_phase_voltage_angles=[]
    bus_C_phase_voltage_angles=[]
    bus_N_voltage_angles=[]
    bus_pos_seq_voltage_magnitudes=[]
    bus_neg_seq_voltage_magnitudes=[]
    bus_avg_voltage_magnitudes=[]
    transformer_names=[]
    transformer_avg_ln_voltage_mags=[]
    transformer_avg_lg_voltage_mags=[]
    transformer_pos_seq_voltage_mags=[]
    transformer_active_power_flows=[]
    transformer_reactive_power_flows=[]
    transformer_TODs=[]
    transformer_loadings=[]
    line_names=[]
    line_levels=[]
    line_TODs=[]
    line_loadings=[]
    substation_names=[]
    line_substations=[]
    line_types=[]

    return bus_levels, bus_TODs, bus_names, bus_A_phase_voltages, bus_B_phase_voltages, bus_C_phase_voltages, bus_N_voltages, bus_A_phase_voltage_angles, bus_B_phase_voltage_angles, bus_C_phase_voltage_angles, bus_N_voltage_angles, bus_pos_seq_voltage_magnitudes, bus_neg_seq_voltage_magnitudes, bus_avg_voltage_magnitudes, substation_names, transformer_names, transformer_avg_ln_voltage_mags, transformer_avg_lg_voltage_mags, transformer_pos_seq_voltage_mags, transformer_active_power_flows, transformer_reactive_power_flows, transformer_TODs, transformer_loadings, line_names, line_levels, line_TODs, line_loadings, line_substations, line_types 


def findMatchingSubstations(subNames, allSubs):
    substationObjects = []
    for idx, subName in enumerate(subNames):
        matchFound = False

        for sub in allSubs:          
            if sub.GetAttribute('loc_name') == subName:
                matchFound = True
                substationObjects.append(sub)
                break

        if not matchFound:
            return substationObjects, 1

    return substationObjects, 0

def checkDataFile(dgApp: object, excelPath: str):
    # Once these are established, load in the data spreadsheet, and do some checks
    # Read in the file from the nominated filepath
    data = read_excel(excelPath, header=None)

    """
    Utilisation of the data is dependent on the file structure
    Columns:
    0: Data & time of day entries
    1: Solar generation
    2: Start of substation specific data
    ----
    n: End of the file, may vary

    Row:
    0: Monitor data file (Y/N) = 
    1: Transformer Number
    2: Load Curve Type = 
    3: Transformer Capacity kVA = 
    4: Transformer peak load kW = 
    5: Total No of homes supplied = 
    6: Solar Inverter Capacity kVA = 
    7: Max Solar Generation kW = 
    8: No of Solar Homes = 
    9: Peak Day PU load Curve TOD
    10: 00:00:00
    ---
    57: 23:30:00

    Column:
    0: Elements above under row values
    1: Solar generation curves
    etc
    """

    substationNames = []
    substationNamesCol = 1
    for col in data.keys().values:
        
        # Skip the desired columns
        if col < 2:
            continue

        substationNames.append(data[col][substationNamesCol])

    # Obtain all the substations in the active study case
    allSubstations = dgApp.GetCalcRelevantObjects('ElmTrfstat')

    # Find the list of matching substations
    matchingSubstations, err = findMatchingSubstations(substationNames, allSubstations)

    # The following will terminate the script if the correct number of matching substations
    # are not found
    if err > 0:
        # dgApp.PrintPlain('Not all the substations in the Excel sheet had matching data in the model')
        # dgApp.PrintPlain('Up to the following substations were matched')
        for match in matchingSubstations:
            dgApp.PrintPlain(match)
        # dgApp.PrintPlain('Please correct the substation names so that they match the model!')
        return matchingSubstations, 1

    return matchingSubstations, 0

def loadSolarProfile(dgApp: object, excelPath: str):
    # Once these are established, load in the data spreadsheet, and do some checks
    # Read in the file from the nominated filepath
    data = read_excel(excelPath, header=None)

    """
    Utilisation of the data is dependent on the file structure
    Columns:
    0: Data & time of day entries
    1: Solar generation
    2: Start of substation specific data
    ----
    n: End of the file, may vary

    Row:
    0: Monitor data file (Y/N) = 
    1: Transformer Number
    2: Load Curve Type = 
    3: Transformer Capacity kVA = 
    4: Transformer peak load kW = 
    5: Total No of homes supplied = 
    6: Solar Inverter Capacity kVA = 
    7: Max Solar Generation kW = 
    8: No of Solar Homes = 
    9: Peak Day PU load Curve TOD
    10: 00:00:00
    ---
    57: 23:30:00

    Column:
    0: Elements above under row values
    1: Solar generation curves
    etc
    """

    solarProfile = []

    for row in range(10, 58):
        solarProfile.append(data[1][row])

    return solarProfile

def loadTodProfile(dgApp: object, excelPath: str):
    # Once these are established, load in the data spreadsheet, and do some checks
    # Read in the file from the nominated filepath
    data = read_excel(excelPath, header=None)

    """
    Utilisation of the data is dependent on the file structure
    Columns:
    0: Data & time of day entries
    1: Solar generation
    2: Start of substation specific data
    ----
    n: End of the file, may vary

    Row:
    0: Monitor data file (Y/N) = 
    1: Transformer Number
    2: Load Curve Type = 
    3: Transformer Capacity kVA = 
    4: Transformer peak load kW = 
    5: Total No of homes supplied = 
    6: Solar Inverter Capacity kVA = 
    7: Max Solar Generation kW = 
    8: No of Solar Homes = 
    9: Peak Day PU load Curve TOD
    10: 00:00:00
    ---
    57: 23:30:00

    Column:
    0: Elements above under row values
    1: Solar generation curves
    etc
    """

    todProfile = []

    # Extract and reformat datetime.time as a string
    for row in range(10, 58):
        todProfile.append(data[0][row].strftime("%H:%M"))

    return todProfile

def loadSubLoadProfile(dgApp: object, excelPath: str):
    # Once these are established, load in the data spreadsheet, and do some checks
    # Read in the file from the nominated filepath
    data = read_excel(excelPath, header=None)

    """
    Utilisation of the data is dependent on the file structure
    Columns:
    0: Data & time of day entries
    1: Solar generation
    2: Start of substation specific data
    ----
    n: End of the file, may vary

    Row:
    0: Monitor data file (Y/N) = 
    1: Transformer Number
    2: Load Curve Type = 
    3: Transformer Capacity kVA = 
    4: Transformer peak load kW = 
    5: Total No of homes supplied = 
    6: Solar Inverter Capacity kVA = 
    7: Max Solar Generation kW = 
    8: No of Solar Homes = 
    9: Peak Day PU load Curve TOD
    10: 00:00:00
    ---
    57: 23:30:00

    Column:
    0: Elements above under row values
    1: Solar generation curves
    etc
    """

    allSubLoads = []

    #substationNames = []
    substationNamesRow = 1
    for col in data.keys().values:

        subload = []

        # Skip the desired columns
        if col < 2:
            continue

        # Removed the name in the first entry
        #subload.append(data[col][substationNamesRow])

        for row in range(10, 58):
            subload.append(data[col][row])

        allSubLoads.append(subload)

    return allSubLoads

def extractObjectsFromSet(dplSet):
    tmpSet = dplSet.GetContents()
    objSet = []
    for tmp in tmpSet:
        objSet.append(tmp.GetAttribute('obj_id'))

    return objSet

def set_loads(all_loads: list, EVs_On: int, EV_map: list, charging_power, tod_load_scaler):

    for idx,obj in enumerate(all_loads):
        name=obj.GetAttribute('loc_name')
        if EVs_On:
            if name[13:14]=="l":
                EV_map_idx=13*(int(name[0:5])/400-1)
            elif name[13:14]=="A":
                EV_map_idx=13*(int(name[0:5])/400-1)+int(name[6:9])/50
            elif name[13:14]=="B":
                EV_map_idx=13*(int(name[0:5])/400-1)+6+int(name[6:9])/50
            EV_count=EV_map[int(EV_map_idx)]
            if name[6:9] in ['300','200','150','050']: # Each PV component represents the PV on the rooftops of 2 houses
                obj.SetAttribute('slinis',EV_count*charging_power/2+2.0*4.008*tod_load_scaler)
                obj.SetAttribute('slinit',EV_count*charging_power/2+2.0*4.008*tod_load_scaler)
            if name[6:9] in ['100','250']: # Each PV component represents the PV on the rooftops of 4 houses
                obj.SetAttribute('slinir',EV_count*charging_power+4.0*4.008*tod_load_scaler)
            if name[6:9]=='000':  # First PV component represents the PV on the rooftops of 2 houses, second two on the rooftops of 1 house each
                obj.SetAttribute('slinir',EV_count*charging_power/2+2.0*4.008*tod_load_scaler)
                obj.SetAttribute('slinis',EV_count*charging_power/4+4.008*tod_load_scaler)
                obj.SetAttribute('slinit',EV_count*charging_power/4+4.008*tod_load_scaler)
        else:
            if name[6:9] in ['300','200','150','050']: # Each PV component represents the PV on the rooftops of 2 houses
                obj.SetAttribute('slinis',2.0*4.008*tod_load_scaler)
                obj.SetAttribute('slinit',2.0*4.008*tod_load_scaler)
            if name[6:9] in ['100','250']: # Each PV component represents the PV on the rooftops of 4 houses
                obj.SetAttribute('slinir',4.0*4.008*tod_load_scaler)
            if name[6:9]=='000':  # First PV component represents the PV on the rooftops of 2 houses, second two on the rooftops of 1 house each
                obj.SetAttribute('slinir',2.0*4.008*tod_load_scaler)
                obj.SetAttribute('slinis',4.008*tod_load_scaler)
                obj.SetAttribute('slinit',4.008*tod_load_scaler)

def save_curtailment_results(app: object, all_QDSLs: list, todIdx: str):
    Perrors=[]
    Qerrors=[]
    voltages=[]
    Q_injections=[]
    curtailments=[]
    utilisation_factors=[]
    Pset_Psum_diffs=[]
    max_P_error=0
    max_Q_error=0
    max_Q_injection=0
    max_curtailment=0
    current_power_outputs=[]
    Buses=[]
    S_values=[]
    P_injections=[]
    PV_list=[]
    for idx, QDSL in enumerate(all_QDSLs):
        if QDSL.GetAttribute('objects')[0].GetAttribute('bus1').GetAttribute('cterm').GetAttribute('uknom')<10 and QDSL.GetAttribute('outserv')==0 and QDSL.GetAttribute('typ_id').GetAttribute('loc_name')=="Volt_Watt_VAr_without_export_limits":
            PV = QDSL.GetAttribute('objects')[0]
            PV_list.append(PV.GetAttribute('loc_name'))
            Buses.append(PV.GetAttribute('bus1').GetAttribute('cterm'))
            if QDSL.GetAttribute('typ_id')=="Volt_Watt_VAr_with_export_limits":
                load = QDSL.GetAttribute('objects')[1]
                Export_limit=QDSL.GetAttribute('Export_limit')
            if PV.GetAttribute('bus1').GetAttribute('nphase')==3:
                PV_phasing=4
            else: 
                PV_phasing=PV.GetAttribute('bus1').GetAttribute('it2p1')
            S=PV.GetAttribute('sgn')
            S_values.append(S)
            if PV_phasing==0:
                voltage=PV.GetAttribute('bus1').GetAttribute('m:uln:A')
                if QDSL.GetAttribute('typ_id')=="Volt_Watt_VAr_with_export_limits":
                    Effective_export_limit=Export_limit+load.GetAttribute('m:P:bus1:A')/QDSL.GetAttribute('Load_division_factor')
                voltages.append(voltage)
            elif PV_phasing==1:
                voltage=PV.GetAttribute('bus1').GetAttribute('m:uln:B')
                if QDSL.GetAttribute('typ_id')=="Volt_Watt_VAr_with_export_limits":
                    Effective_export_limit=Export_limit+load.GetAttribute('m:P:bus1:B')/QDSL.GetAttribute('Load_division_factor')
                voltages.append(voltage)
            elif PV_phasing==2:
                voltage=PV.GetAttribute('bus1').GetAttribute('m:uln:C')
                if QDSL.GetAttribute('typ_id')=="Volt_Watt_VAr_with_export_limits":
                    Effective_export_limit=Export_limit+load.GetAttribute('m:P:bus1:C')/QDSL.GetAttribute('Load_division_factor')
                voltages.append(voltage)
            elif PV_phasing==4:
                voltage=PV.GetAttribute('bus1').GetAttribute('m:u1')
                if QDSL.GetAttribute('typ_id')=="Volt_Watt_VAr_with_export_limits":
                    Effective_export_limit=Export_limit+(load.GetAttribute('m:P:bus1:A')+load.GetAttribute('m:P:bus1:B')+load.GetAttribute('m:P:bus1:C'))/QDSL.GetAttribute('Load_division_factor')
                voltages.append(voltage)
            Pset=PV.GetAttribute('s:pset')*1000
            Qset=PV.GetAttribute('s:qset')*1000
            if QDSL.GetAttribute('volt_var_active'):
                if voltage<0.88026:               
                    if voltage>0.828024:
                        Q_target=0.44*S*(0.88026-voltage)/(0.88026-0.828024)
                    else:
                        Q_target=0.44*S
                elif voltage>0.960028161:
                    if voltage<1.032030273:
                        Q_target=-0.6*S*(voltage-0.960028161)/(1.032030273-0.960028161)
                    else:
                        Q_target=-0.6*S
                else:
                    Q_target=0
                
                if abs(Q_target)>0:
                    if Pset<(0.1*S):
                        Q_limiter=Pset*6*Q_target/abs(Q_target)
                    else:
                        Q_limiter=S*Q_target/abs(Q_target)
                    Q_target=min(abs(Q_target),abs(Q_limiter))*abs(Q_target)/Q_target                        
            else:
                Q_target=0
            if voltage>1.01203:
                if voltage<1.040031:
                    P_VW_limiter=0.8*S*(1.040031-voltage)/(1.040031-1.01203)+0.2*S
                else:
                    P_VW_limiter=0.2*S
            else:
                P_VW_limiter=S
            P_rating_limiter=(S**2-Q_target**2)**0.5
            pgini=PV.GetAttribute('pgini')
            if QDSL.GetAttribute('typ_id')=="Volt_Watt_VAr_with_export_limits":
                Ptarget=min(P_VW_limiter,P_rating_limiter,pgini,Effective_export_limit)
            else:
                Ptarget=min(P_VW_limiter,P_rating_limiter,pgini)
            Perror=(Pset-Ptarget)/S*100
            if abs(Q_target)>0:
                Qerror=(Qset-Q_target)/S*100
            else:
                Qerror=0
            curtailment=(pgini-Pset)/S*100
            if curtailment>max_curtailment:
                max_curtailment=curtailment
            if abs(Q_target)>max_Q_injection:
                max_Q_injection=abs(Q_target)
            if abs(Perror)>max_P_error:
                max_P_error=Perror
            Perrors.append(Perror)
            if abs(Qerror)>max_Q_error:
                max_Q_error=Qerror
            Qerrors.append(abs(Qerror))
            curtailments.append(curtailment)
            P_injections.append(Pset/S*100)
            Q_injections.append(Qset/S*100)
            current_power_outputs.append(Pset)
            utilisation_factor=(Pset**2+Qset**2)**0.5/S*100
            utilisation_factors.append(utilisation_factor)
            Pset_Psum_diffs.append(Pset-PV.GetAttribute('m:Psum:bus1'))
    wb = Workbook()
    ws = wb.active
    for idx, item in enumerate(PV_list):
        row=[]
        if idx==0:
            row.append('PV Names')
            row.append('Bus Names')
            row.append('Rating (kVA)')
            row.append('Curtailments (%)')
            row.append('P injections (%)')
            row.append('Q injections (%)')
            row.append("P error (%)")
            row.append("Q error (%)")
            row.append("Substation Name")
            ws.append(row)
            row=[]
        row.append(item)
        row.append(Buses[idx].GetAttribute('loc_name'))
        row.append(S_values[idx])
        row.append(curtailments[idx])
        row.append(P_injections[idx])
        row.append(Q_injections[idx])
        row.append(Perrors[idx])
        row.append(Qerrors[idx])
        try:
            Bus=Buses[idx]              
            row.append(Bus.GetAttribute('cpSubstat').GetAttribute('loc_name'))
        except:
            row.append(" ")
        ws.append(row)
    wb.save('Curtailment_results_'+todIdx+'.xlsx')

def pack_bus_results(QDSLsOn, solarOn, Season, app, todIdx: int, bus_levels: list, bus_names: list, bus_TODs: list, bus_A_phase_voltages: list, bus_B_phase_voltages: list, bus_C_phase_voltages: list, bus_N_voltages: list, bus_A_phase_voltage_angles: list, bus_B_phase_voltage_angles: list, bus_C_phase_voltage_angles: list,bus_N_voltage_angles: list, bus_pos_seq_voltage_magnitudes: list, bus_neg_seq_voltage_magnitudes: list, bus_avg_voltage_magnitudes: list, all_buses: list, tod: str, substation_names: list):
    
    for bus in all_buses:
        if "dummy" not in bus.GetAttribute('loc_name'):
            if bus.GetAttribute('uknom')<10:
                bus_levels.append("LV")
            else:
                bus_levels.append("MV")
            bus_names.append(bus.GetAttribute('loc_name'))
            bus_TODs.append(tod)
            bus_A_phase_voltages.append(1000*bus.GetAttribute('m:Uln:A'))
            bus_B_phase_voltages.append(1000*bus.GetAttribute('m:Uln:B'))
            bus_C_phase_voltages.append(1000*bus.GetAttribute('m:Uln:C'))
            bus_A_phase_voltage_angles.append(bus.GetAttribute('m:phiuln:A'))
            bus_B_phase_voltage_angles.append(bus.GetAttribute('m:phiuln:B'))
            bus_C_phase_voltage_angles.append(bus.GetAttribute('m:phiuln:C'))
            bus_N_voltages.append(1000*bus.GetAttribute('m:Un'))
            bus_N_voltage_angles.append(bus.GetAttribute('m:phiun'))
            bus_pos_seq_voltage_magnitudes.append(1000*bus.GetAttribute('m:U1'))
            bus_neg_seq_voltage_magnitudes.append(1000*bus.GetAttribute('m:U2'))
            bus_avg_voltage_magnitudes.append(1000*bus.GetAttribute('m:Um'))
            try:
                substation_names.append(bus.GetAttribute('cpSubstat').GetAttribute('loc_name'))
            except: 
                substation_names.append(" ")                

    return bus_levels, bus_names, bus_TODs, bus_A_phase_voltages, bus_B_phase_voltages, bus_C_phase_voltages, bus_N_voltages, bus_A_phase_voltage_angles, bus_B_phase_voltage_angles, bus_C_phase_voltage_angles, bus_N_voltage_angles, bus_pos_seq_voltage_magnitudes, bus_neg_seq_voltage_magnitudes,  bus_avg_voltage_magnitudes, substation_names

def pack_transformer_results(transformer_names: list, transformer_TODs: list, transformer_avg_ln_voltage_mags: list, transformer_active_power_flows: list, transformer_reactive_power_flows: list, transformer_loadings: list, all_transformers: list, tod: str):
    
    for transformer in all_transformers:
        transformer_names.append(transformer.GetAttribute('loc_name'))           
        transformer_TODs.append(tod)
        vlnA=transformer.GetAttribute('buslv').GetAttribute('cterm').GetAttribute('m:Uln:A')*1000
        vlnB=transformer.GetAttribute('buslv').GetAttribute('cterm').GetAttribute('m:Uln:B')*1000
        vlnC=transformer.GetAttribute('buslv').GetAttribute('cterm').GetAttribute('m:Uln:C')*1000
        transformer_avg_ln_voltage_mags.append((vlnA+vlnB+vlnC)/3)
        transformer_active_power_flows.append(transformer.GetAttribute('m:Psum:bushv'))
        transformer_reactive_power_flows.append(transformer.GetAttribute('m:Qsum:bushv'))
        transformer_loadings.append(transformer.GetAttribute('c:loading'))

    return transformer_names, transformer_TODs, transformer_avg_ln_voltage_mags, transformer_active_power_flows, transformer_reactive_power_flows, transformer_loadings

def pack_line_results(line_names: list, line_TODs: list, line_levels: list, line_loadings: list,all_lines: list, tod: str, line_substations: list, line_types: list):

    for line in all_lines:
        if "dummy" not in line.GetAttribute('loc_name'):
            line_names.append(line.GetAttribute('loc_name'))
            line_TODs.append(tod)
            if line.GetAttribute('bus1').GetAttribute('cterm').GetAttribute('uknom')>10:
                line_levels.append("MV")
            else:
                line_levels.append("LV")
            line_loadings.append(line.GetAttribute('c:loading'))
            try:
                line_substations.append(line.GetAttribute('bus1').GetAttribute('cterm').GetAttribute('cpSubstat').GetAttribute('loc_name'))
            except:
                line_substations.append(" ")
            try:
                if "OHL" in line.GetAttribute('c_ptow').GetAttribute('loc_name'):
                    line_types.append("OH")
                else:
                    line_types.append("UG")
            except:
                line_types.append(" ")

    return line_names, line_TODs, line_levels, line_loadings, line_substations, line_types

def save_bus_results(todIdx, bus_levels, bus_TODs, QDSLsOn, solarOn, bus_names, bus_A_phase_voltages, bus_B_phase_voltages, bus_C_phase_voltages, bus_N_voltages, bus_A_phase_voltage_angles, bus_B_phase_voltage_angles, bus_C_phase_voltage_angles, bus_N_voltage_angles, bus_pos_seq_voltage_magnitudes, bus_neg_seq_voltage_magnitudes, bus_avg_voltage_magnitudes, Season, app, substation_names):

    wb = Workbook()
    ws = wb.active
    for idx, item in enumerate(bus_levels):
        # app.PrintPlain(idx/len(bus_levels))
        row=[]
        if idx==0:
            row.append('TOD')
            row.append('QDSLs')
            row.append('Solar')
            row.append("Bus Voltage Levels")
            row.append("Bus Names")
            row.append("Bus A Phase Voltages (V)")
            row.append("Bus B Phase Voltages (V)")
            row.append("Bus C Phase Voltages (V)")
            row.append("Bus Neutral Voltages (V)") 
            row.append("Bus A Phase Voltage Angles (deg)")
            row.append("Bus B Phase Voltage Angles (deg)")
            row.append("Bus C Phase Voltage Angles (deg)")
            row.append("Bus Neutral Voltage Angles (deg)")                
            row.append("Bus Positive Sequence Voltage Magnitudes (V)")
            row.append("Bus Negative Sequence Voltage Magnitudes (V)")
            row.append("Bus Average Voltage Magnitudes (V)")
            row.append("Substation Name")
            ws.append(row)
            row=[]
        row.append(bus_TODs[idx])
        if not(QDSLsOn):
            row.append("Off")
        else:
            row.append("On")
        if not(solarOn):
            row.append("Off")
        else:
            row.append("On")
        row.append(item)
        row.append(bus_names[idx])
        row.append(bus_A_phase_voltages[idx])
        row.append(bus_B_phase_voltages[idx])
        row.append(bus_C_phase_voltages[idx])
        row.append(bus_N_voltages[idx])
        row.append(bus_A_phase_voltage_angles[idx])
        row.append(bus_B_phase_voltage_angles[idx])
        row.append(bus_C_phase_voltage_angles[idx])
        row.append(bus_N_voltage_angles[idx])
        row.append(bus_pos_seq_voltage_magnitudes[idx])
        row.append(bus_neg_seq_voltage_magnitudes[idx])
        row.append(bus_avg_voltage_magnitudes[idx])
        row.append(substation_names[idx])
        ws.append(row)
    if solarOn:
        if QDSLsOn:
            wb.save('Bus_voltage_results_QDSLsOn_his_'+Season+"_"+"_"+str(todIdx)+'.xlsx')
        else:
            wb.save('Bus_voltage_results_QDSLsOff_his_'+Season+"_"+"_"+str(todIdx)+'.xlsx')
    else:
        if QDSLsOn:
            wb.save('Bus_voltage_results_QDSLsOn_los_'+Season+"_"+"_"+str(todIdx)+'.xlsx')
        else:
            wb.save('Bus_voltage_results_QDSLsOff_los_'+Season+"_"+"_"+str(todIdx)+'.xlsx')
    
def save_transformer_results(transformer_names: list, app: object, transformer_TODs: list, QDSLsOn: int, solarOn: int, transformer_avg_ln_voltage_mags: list, transformer_active_power_flows: list, transformer_reactive_power_flows: list, transformer_loadings: list, Season: str):

    wb = Workbook()
    ws = wb.active 
    for idx, item in enumerate(transformer_names):
        # app.PrintPlain(idx/len(transformer_names))
        row=[]
        if idx==0:
            row.append('TOD')
            row.append('QDSLs')
            row.append('Solar')
            row.append("Transformer Names")
            row.append("Transformer Average Line-Neutral Voltage Magnitudes (V)")
            # row.append("Transformer Average Line-Ground Voltage Magnitudes (V)")
            # row.append("Transformer Positive Sequence Voltage Magnitudes (V)")
            row.append("Transformer Active Power Flows (kW)")
            row.append("Transformer Reactive Power Flows (kW)")
            row.append("Loading (%)")
            ws.append(row)
            row=[]
        row.append(transformer_TODs[idx])
        if not(QDSLsOn):
            row.append("Off")
        else:
            row.append("On")
        if not(solarOn):
            row.append("Off")
        else:
            row.append("On")
        row.append(item)
        row.append(transformer_avg_ln_voltage_mags[idx])
        # row.append(transformer_avg_lg_voltage_mags[idx])
        # row.append(transformer_pos_seq_voltage_mags[idx])
        row.append(transformer_active_power_flows[idx])
        row.append(transformer_reactive_power_flows[idx])
        row.append(transformer_loadings[idx])
        ws.append(row)
    if solarOn:
        if QDSLsOn:
            wb.save('Transformer_results_QDSLsOn_his_'+Season+'.xlsx')
        else:
            wb.save('Transformer_results_QDSLsOff_his_'+Season+'.xlsx')
    else:
        if QDSLsOn:
            wb.save('Transformer_results_QDSLsOn_los_'+Season+'.xlsx')
        else:
            wb.save('Transformer_results_QDSLsOff_los_'+Season+'.xlsx')

def save_line_results(app: object, line_TODs: list, QDSLsOn: int, solarOn: int, line_levels: list, line_names: list, line_loadings: list, Season: str, line_substations: list, line_types: list):
        
    wb = Workbook()
    ws = wb.active
    for idx, item in enumerate(line_TODs):
        # app.PrintPlain(idx/len(line_TODs))
        row=[]
        if idx==0:
            row.append('TOD')
            row.append('QDSLs')
            row.append('Solar')
            row.append("Line Voltage Levels")
            row.append("Line Name")
            row.append("Loading (%)")
            row.append("Substion Name")
            row.append("Line Type (UG or OH)")
            ws.append(row)
            row=[]
        row.append(line_TODs[idx])
        if not(QDSLsOn):
            row.append("Off")
        else:
            row.append("On")
        if not(solarOn):
            row.append("Off")
        else:
            row.append("On")       
        row.append(line_levels[idx])
        row.append(line_names[idx])
        row.append(line_loadings[idx])
        row.append(line_substations[idx])
        row.append(line_types[idx])
        ws.append(row)
    if solarOn:
        if QDSLsOn:
            wb.save('Line_loading_results_QDSLsOn_his_'+Season+'.xlsx')
        else:
            wb.save('Line_loading_results_QDSLsOff_his_'+Season+'.xlsx')
    else:
        if QDSLsOn:
            wb.save('Line_loading_results_QDSLsOn_los_'+Season+'.xlsx')
        else:
            wb.save('Line_loading_results_QDSLsOff_los_'+Season+'.xlsx')
  
def construct_EV_maps(all_LV_loads,app):

    charger_size_1_list=['a','b','c','d','e','f','A','B','C','D','E','F']
    charger_size_2_list=['3','T']
    total_size1_charger_EVs=0
    total_size2_charger_EVs=0
    EVs_to_loads_map1=[]
    EVs_to_phases_map1=[]
    EVs_to_loads_map2=[]
    for load in all_LV_loads:
        load_name=load.GetAttribute('loc_name')
        if "bal_" in load_name:
            EV_substring=load_name[load_name.find("bal_")+6:len(load_name)]
            local_no_EVs=len(load_name)-load_name.find("Unbal_")-8
            for i in range(local_no_EVs):
                if EV_substring[i] in charger_size_1_list:
                    EVs_to_loads_map1.append(load)
                    EVs_to_phases_map1.append(load_name[load_name.find("bal_")+6+i])
                elif EV_substring[i] in charger_size_2_list:
                    EVs_to_loads_map2.append(load)
        else:
            EV_substring=load_name[load_name.find("al")+2:len(load_name)]
            local_no_EVs=len(load_name)-load_name.find("al")-2
            for i in range(local_no_EVs):
                if EV_substring[i] in charger_size_1_list:
                    EVs_to_loads_map1.append(load)
                    EVs_to_phases_map1.append(load_name[load_name.find("al")+2+i])
                elif EV_substring[i] in charger_size_2_list:
                    EVs_to_loads_map2.append(load)
        for i in EV_substring:
            if i in charger_size_1_list:
                total_size1_charger_EVs=total_size1_charger_EVs+1
            elif i in charger_size_2_list:
                total_size2_charger_EVs=total_size2_charger_EVs+1   

    app.PrintPlain("Total no of EVs with charger size 1 is "+str(total_size1_charger_EVs))
    app.PrintPlain("Total no of EVs with charger size 2 is "+str(total_size2_charger_EVs))

    return EVs_to_loads_map1, EVs_to_phases_map1, EVs_to_loads_map2, total_size1_charger_EVs, total_size2_charger_EVs

def activate_EV(load,phase_char,app):
    load_name=load.GetAttribute('loc_name')
    if "bal_" in load_name:
        EV_substring=load_name[load_name.find("bal_")+6:len(load_name)]
    else:
        EV_substring=load_name[load_name.find("al")+2:len(load_name)]
    idx=EV_substring.find(phase_char)+len(load_name)-len(EV_substring)
    if phase_char=="a":
        new_load_name=load_name[0:idx]+"A"+load_name[idx+1:len(load_name)]
    elif phase_char=="b":
        new_load_name=load_name[0:idx]+"B"+load_name[idx+1:len(load_name)]
    elif phase_char=="c":
        new_load_name=load_name[0:idx]+"C"+load_name[idx+1:len(load_name)]
    elif phase_char=="3":
        new_load_name=load_name[0:idx]+"T"+load_name[idx+1:len(load_name)]
    elif phase_char=="d":
        new_load_name=load_name[0:idx]+"D"+load_name[idx+1:len(load_name)]
    elif phase_char=="e":
        new_load_name=load_name[0:idx]+"E"+load_name[idx+1:len(load_name)]
    elif phase_char=="f":
        new_load_name=load_name[0:idx]+"F"+load_name[idx+1:len(load_name)]
    load.SetAttribute("loc_name",new_load_name)
    
def deactivate_EV(load,phase_char,app):
    load_name=load.GetAttribute('loc_name')
    # app.PrintPlain("New load...")
    # app.PrintPlain(load_name)
    if "bal_" in load_name:
        EV_substring=load_name[load_name.find("bal_")+6:len(load_name)]
    else:
        EV_substring=load_name[load_name.find("al")+2:len(load_name)]
    idx=EV_substring.find(phase_char)+len(load_name)-len(EV_substring)
    if phase_char=="A":
        new_load_name=load_name[0:idx]+"a"+load_name[idx+1:len(load_name)]
    elif phase_char=="B":
        new_load_name=load_name[0:idx]+"b"+load_name[idx+1:len(load_name)]
    elif phase_char=="C":
        new_load_name=load_name[0:idx]+"c"+load_name[idx+1:len(load_name)]
    elif phase_char=="T":
        new_load_name=load_name[0:idx]+"3"+load_name[idx+1:len(load_name)]
    elif phase_char=="D":
        new_load_name=load_name[0:idx]+"d"+load_name[idx+1:len(load_name)]
    elif phase_char=="E":
        new_load_name=load_name[0:idx]+"e"+load_name[idx+1:len(load_name)]
    elif phase_char=="F":
        new_load_name=load_name[0:idx]+"f"+load_name[idx+1:len(load_name)]
    load.SetAttribute("loc_name",new_load_name)
    # load_name=load.GetAttribute('loc_name')
    # app.PrintPlain(load_name)
    # app.PrintPlain(" ")
    
def set_EVs_charging(EVs_charging_list1, EVs_charging_list2, EVs_to_loads_map1, EVs_to_phases_map1, EVs_to_loads_map2, total_size1_charger_EVs, total_size2_charger_EVs, EV_curve1, EV_curve2, todIdx, current_EV_charging_count1, current_EV_charging_count2, app):

    target_EV_charging_proportion1=EV_curve1[todIdx]
    target_EV_charging_proportion2=EV_curve2[todIdx]
    # app.PrintPlain(target_EV_charging_proportion)
    current_EV_charging_proportion1=current_EV_charging_count1/total_size1_charger_EVs
    current_EV_charging_proportion2=current_EV_charging_count2/total_size2_charger_EVs
    

    if current_EV_charging_proportion1<target_EV_charging_proportion1:
        while current_EV_charging_proportion1<target_EV_charging_proportion1:
            x=round(random.uniform(0,len(EVs_to_loads_map1)-1))
            load=EVs_to_loads_map1[x]
            if x not in EVs_charging_list1:
                EVs_charging_list1.append(x)
                rating=load.GetAttribute('slini')
                current_scaling_factor=load.GetAttribute('scale0')
                current_effective_size=rating*current_scaling_factor
                new_scaling_factor=(current_effective_size+3.68)/(rating+3.68) # Sum of what we want it to be currently drawing over sum of what we want it to be theoretically capable of drawing
                load.SetAttribute('scale0',new_scaling_factor)
                if EVs_to_phases_map1[x]=="a":
                    load.SetAttribute('slinir',load.GetAttribute('slinir')+3.68)
                    activate_EV(load,"a",app)
                elif EVs_to_phases_map1[x]=="b":
                    load.SetAttribute('slinis',load.GetAttribute('slinis')+3.68)
                    activate_EV(load,"b",app)
                elif EVs_to_phases_map1[x]=="c":
                    load.SetAttribute('slinit',load.GetAttribute('slinit')+3.68)
                    activate_EV(load,"c",app)
                elif EVs_to_phases_map1[x]=="d":
                    load.SetAttribute('slinir',load.GetAttribute('slinir')+3.68/2)
                    load.SetAttribute('slinis',load.GetAttribute('slinis')+3.68/2)
                    activate_EV(load,"d",app)
                elif EVs_to_phases_map1[x]=="e":
                    load.SetAttribute('slinir',load.GetAttribute('slinir')+3.68/2)
                    load.SetAttribute('slinit',load.GetAttribute('slinit')+3.68/2)
                    activate_EV(load,"e",app)
                elif EVs_to_phases_map1[x]=="f":
                    load.SetAttribute('slinis',load.GetAttribute('slinis')+3.68/2)
                    load.SetAttribute('slinit',load.GetAttribute('slinit')+3.68/2)
                    activate_EV(load,"f",app)
                current_EV_charging_count1=current_EV_charging_count1+1
                current_EV_charging_proportion1=current_EV_charging_count1/total_size1_charger_EVs
    else:    
        while current_EV_charging_proportion1>target_EV_charging_proportion1:
            x=round(random.uniform(0,len(EVs_to_loads_map1)-1))
            load=EVs_to_loads_map1[x]
            if x in EVs_charging_list1:
                EVs_charging_list1.remove(x)
                rating=load.GetAttribute('slini')
                current_scaling_factor=load.GetAttribute('scale0')
                current_effective_size=rating*current_scaling_factor
                new_scaling_factor=(current_effective_size-3.68)/(rating-3.68) # Sum of what we want it to be currently drawing over sum of what we want it to be theoretically capable of drawing
                load.SetAttribute('scale0',new_scaling_factor)
                if EVs_to_phases_map1[x]=="a":
                    load.SetAttribute('slinir',load.GetAttribute('slinir')-3.68)
                    deactivate_EV(load,"A",app)
                elif EVs_to_phases_map1[x]=="b":
                    load.SetAttribute('slinis',load.GetAttribute('slinis')-3.68)
                    deactivate_EV(load,"B",app)
                elif EVs_to_phases_map1[x]=="c":
                    load.SetAttribute('slinit',load.GetAttribute('slinit')-3.68)
                    deactivate_EV(load,"C",app)
                elif EVs_to_phases_map1[x]=="d":
                    load.SetAttribute('slinir',load.GetAttribute('slinir')-3.68/2)
                    load.SetAttribute('slinis',load.GetAttribute('slinis')-3.68/2)
                    deactivate_EV(load,"D",app)
                elif EVs_to_phases_map1[x]=="e":
                    load.SetAttribute('slinir',load.GetAttribute('slinir')-3.68/2)
                    load.SetAttribute('slinit',load.GetAttribute('slinit')-3.68/2)
                    deactivate_EV(load,"E",app)
                elif EVs_to_phases_map1[x]=="f":
                    load.SetAttribute('slinis',load.GetAttribute('slinis')-3.68/2)
                    load.SetAttribute('slinit',load.GetAttribute('slinit')-3.68/2)
                    deactivate_EV(load,"F",app)
                current_EV_charging_count1=current_EV_charging_count1-1
                current_EV_charging_proportion1=current_EV_charging_count1/total_size1_charger_EVs
    
    if current_EV_charging_proportion2<target_EV_charging_proportion2:
        while current_EV_charging_proportion2<target_EV_charging_proportion2:
            x=round(random.uniform(0,len(EVs_to_loads_map2)-1))
            load=EVs_to_loads_map2[x]
            if x not in EVs_charging_list2:
                EVs_charging_list2.append(x)
                rating=load.GetAttribute('slini')
                current_scaling_factor=load.GetAttribute('scale0')
                current_effective_size=rating*current_scaling_factor
                new_scaling_factor=(current_effective_size+7.36)/(rating+7.36) # Sum of what we want it to be currently drawing over sum of what we want it to be theoretically capable of drawing
                load.SetAttribute('scale0',new_scaling_factor)
                load.SetAttribute('slinir',load.GetAttribute('slini')+7.36)
                activate_EV(load,"3",app)
                current_EV_charging_count2=current_EV_charging_count2+1
                current_EV_charging_proportion2=current_EV_charging_count2/total_size2_charger_EVs
    else:    
        while current_EV_charging_proportion2>target_EV_charging_proportion2:
            x=round(random.uniform(0,len(EVs_to_loads_map2)-1))
            load=EVs_to_loads_map2[x]
            if x in EVs_charging_list2:
                EVs_charging_list2.remove(x)
                rating=load.GetAttribute('slini')
                current_scaling_factor=load.GetAttribute('scale0')
                current_effective_size=rating*current_scaling_factor
                new_scaling_factor=(current_effective_size-7.36)/(rating-7.36) # Sum of what we want it to be currently drawing over sum of what we want it to be theoretically capable of drawing
                load.SetAttribute('scale0',new_scaling_factor)
                load.SetAttribute('slinir',load.GetAttribute('slinir')-7.36)
                deactivate_EV(load,"T",app)
                current_EV_charging_count2=current_EV_charging_count2-1
                current_EV_charging_proportion2=current_EV_charging_count2/total_size2_charger_EVs

    return EVs_charging_list1, EVs_charging_list2, current_EV_charging_count1, current_EV_charging_count2

def deactivate_all_EVs(EVs_charging_list1, EVs_charging_list2,EVs_to_loads_map1,EVs_to_phases_map1,EVs_to_loads_map2,app):
    
    for x in EVs_charging_list1:
        load=EVs_to_loads_map1[x]
        rating=load.GetAttribute('slini')
        current_scaling_factor=load.GetAttribute('scale0')
        current_effective_size=rating*current_scaling_factor
        new_scaling_factor=(current_effective_size-3.68)/(rating-3.68) # Sum of what we want it to be currently drawing over sum of what we want it to be theoretically capable of drawing
        load.SetAttribute('scale0',new_scaling_factor)
        if EVs_to_phases_map1[x]=="a":               
            load.SetAttribute('slinir',load.GetAttribute('slinir')-3.68)
            deactivate_EV(load,"A",app)
        elif EVs_to_phases_map1[x]=="b":
            load.SetAttribute('slinis',load.GetAttribute('slinis')-3.68)
            deactivate_EV(load,"B",app)
        elif EVs_to_phases_map1[x]=="c":                
            load.SetAttribute('slinit',load.GetAttribute('slinit')-3.68)
            deactivate_EV(load,"C",app)           
        elif EVs_to_phases_map1[x]=="d":                
            load.SetAttribute('slinir',load.GetAttribute('slinir')-3.68/2)
            load.SetAttribute('slinis',load.GetAttribute('slinis')-3.68/2)
            deactivate_EV(load,"D",app)
        elif EVs_to_phases_map1[x]=="e":                
            load.SetAttribute('slinir',load.GetAttribute('slinir')-3.68/2)
            load.SetAttribute('slinit',load.GetAttribute('slinit')-3.68/2)
            deactivate_EV(load,"E",app)
        elif EVs_to_phases_map1[x]=="f":                
            load.SetAttribute('slinis',load.GetAttribute('slinis')-3.68/2)
            load.SetAttribute('slinit',load.GetAttribute('slinit')-3.68/2)
            deactivate_EV(load,"F",app)

    for x in EVs_charging_list2:
        load=EVs_to_loads_map2[x]
        rating=load.GetAttribute('slini')
        current_scaling_factor=load.GetAttribute('scale0')
        current_effective_size=rating*current_scaling_factor
        new_scaling_factor=(current_effective_size-7.36)/(rating-7.36) # Sum of what we want it to be currently drawing over sum of what we want it to be theoretically capable of drawing
        load.SetAttribute('scale0',new_scaling_factor)
        load.SetAttribute('slini',load.GetAttribute('slini')-7.36)
        deactivate_EV(load,"T",app)

def rescale_non_EV_LV_load_portion(obj,subLoadScaler,total_size1_EVs, total_size2_EVs, total_size1_EVs_On, total_size2_EVs_On,app):
    
    smaller_charger_size_list=["a","b","c","d","e","f","A","B","C","D","E","F",]
    larger_charger_size_list=["3","T"]
    rating=obj.GetAttribute('slini')
    current_scaling_factor=obj.GetAttribute('scale0')
    load_name=obj.GetAttribute('loc_name')
    if "bal_" in load_name:
        EV_substring=load_name[load_name.find("bal_")+6:len(load_name)]
    else:
        EV_substring=load_name[load_name.find("al")+2:len(load_name)]
    EV_size=0
    for i in EV_substring:
        if i in larger_charger_size_list:
            total_size2_EVs=total_size2_EVs+1
            if i.isupper():        
                EV_size=EV_size+7.36
                total_size2_EVs_On=total_size2_EVs_On+1
        elif i in smaller_charger_size_list:
            total_size1_EVs=total_size1_EVs+1
            if i.isupper():        
                EV_size=EV_size+3.68
                total_size1_EVs_On=total_size1_EVs_On+1
    non_EV_rating_portion=rating-EV_size
    target_effective_non_EV_size=non_EV_rating_portion*subLoadScaler
    new_scaling_factor=(target_effective_non_EV_size+EV_size)/rating
    obj.SetAttribute('scale0', new_scaling_factor)                    

    return total_size1_EVs, total_size2_EVs, total_size1_EVs_On, total_size2_EVs_On

def realise_STATCOMs(all_buses,bus_list,grid,ldf):
    all_STATCOMs=[]
    ldf.SetAttribute('iopt_lim',1)
    for target_bus_name in bus_list:
        for bus in all_buses:
            if bus.GetAttribute('uknom')<10 and bus.GetAttribute('loc_name')==target_bus_name:
                new_a_phase_STATCOM = grid.CreateObject('ElmGenstat', bus.GetAttribute('loc_name')+"_STATCOM_A")
                new_a_phase_STATCOM.SetAttribute('outserv',0)
                new_a_phase_STATCOM.SetAttribute('usetp',0.92)
                new_a_phase_STATCOM.SetAttribute('ddroop',0.5)
                new_a_phase_STATCOM.SetAttribute('usp_min',0.9)
                new_a_phase_STATCOM.SetAttribute('usp_max',0.94)
                new_a_phase_STATCOM.SetAttribute('sgn',0.01666667)     
                new_a_phase_STATCOM.SetAttribute('phtech',3)     
                cubic = bus.CreateObject('StaCubic')
                cubic.SetAttribute('it2p1',0)
                new_a_phase_STATCOM.SetAttribute('bus1', cubic)
                new_a_phase_STATCOM.SetAttribute('cosn', 0.000001)
                new_a_phase_STATCOM.SetAttribute('av_mode', 'vdroop')
                new_a_phase_STATCOM.SetAttribute('Pmax_uc', 0)
                new_a_phase_STATCOM.SetAttribute('P_max', 0)
                
                new_b_phase_STATCOM = grid.CreateObject('ElmGenstat', bus.GetAttribute('loc_name')+"_STATCOM_B")
                new_b_phase_STATCOM.SetAttribute('outserv',0)
                new_b_phase_STATCOM.SetAttribute('usetp',0.92)
                new_b_phase_STATCOM.SetAttribute('ddroop',0.5)
                new_b_phase_STATCOM.SetAttribute('usp_min',0.9)
                new_b_phase_STATCOM.SetAttribute('usp_max',0.94)
                new_b_phase_STATCOM.SetAttribute('sgn',0.01666667)     
                new_b_phase_STATCOM.SetAttribute('phtech',3)     
                cubic = bus.CreateObject('StaCubic')
                cubic.SetAttribute('it2p1',1)
                new_b_phase_STATCOM.SetAttribute('bus1', cubic)
                new_b_phase_STATCOM.SetAttribute('cosn', 0.000001)
                new_b_phase_STATCOM.SetAttribute('av_mode', 'vdroop')
                new_b_phase_STATCOM.SetAttribute('Pmax_uc', 0)
                new_b_phase_STATCOM.SetAttribute('P_max', 0)
                
                new_c_phase_STATCOM = grid.CreateObject('ElmGenstat', bus.GetAttribute('loc_name')+"_STATCOM_C")
                new_c_phase_STATCOM.SetAttribute('outserv',0)
                new_c_phase_STATCOM.SetAttribute('usetp',0.92)
                new_c_phase_STATCOM.SetAttribute('ddroop',0.5)
                new_c_phase_STATCOM.SetAttribute('usp_min',0.9)
                new_c_phase_STATCOM.SetAttribute('usp_max',0.94)
                new_c_phase_STATCOM.SetAttribute('sgn',0.01666667)     
                new_c_phase_STATCOM.SetAttribute('phtech',3)     
                cubic = bus.CreateObject('StaCubic')
                cubic.SetAttribute('it2p1',2)
                new_c_phase_STATCOM.SetAttribute('bus1', cubic)
                new_c_phase_STATCOM.SetAttribute('cosn', 0.000001)
                new_c_phase_STATCOM.SetAttribute('av_mode', 'vdroop')
                new_c_phase_STATCOM.SetAttribute('Pmax_uc', 0)
                new_c_phase_STATCOM.SetAttribute('P_max', 0)

                all_STATCOMs.append(new_a_phase_STATCOM,new_b_phase_STATCOM, new_c_phase_STATCOM)

    return all_STATCOMs
